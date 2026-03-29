from fastapi import APIRouter, HTTPException, Header, UploadFile, File
from pymongo import MongoClient
from datetime import datetime, timedelta, time
from utils.jwt import verify_token
from utils.ai import calculate_wait_time
from bson import ObjectId
from openpyxl import load_workbook
from io import BytesIO
import os

router = APIRouter()
client = MongoClient(os.getenv("MONGO_URL", "mongodb://localhost:27017"))
db = client["synckiet"]

TIME_SLOTS = {
    1: {"start": "09:10", "end": "10:00"},
    2: {"start": "10:00", "end": "10:50"},
    3: {"start": "10:50", "end": "11:40"},
    4: {"start": "11:40", "end": "12:30"},
    5: {"start": "13:30", "end": "14:20"},
    6: {"start": "14:20", "end": "15:10"},
    7: {"start": "15:10", "end": "16:00"},
    8: {"start": "16:00", "end": "16:50"},
}

DAY_MAP = {
    0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday",
    4: "Friday", 5: "Saturday", 6: "Sunday"
}

WORKING_DAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday"}
VALID_DAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}
DEFAULT_RESOLUTION_MINUTES = 15


def get_current_day():
    return DAY_MAP.get(datetime.now().weekday(), "Monday")


def get_current_period():
    now = datetime.now().strftime("%H:%M")
    for period, times in TIME_SLOTS.items():
        if times["start"] <= now <= times["end"]:
            return period
    return None


def verify_user(authorization: str, allowed_roles):
    token = authorization.replace("Bearer ", "")
    user = verify_token(token)
    if not user or user["role"] not in allowed_roles:
        raise HTTPException(401, "Unauthorized")
    return user


def build_timetable_from_slots(slots):
    timetable = {}
    for slot in slots:
        day = slot["day"]
        period = str(slot["period"])
        timetable.setdefault(day, {})
        timetable[day][period] = {
            "subject": slot.get("subject", ""),
            "section": slot.get("section", ""),
            "type": slot.get("class_type", "theory"),
            "room": slot.get("room", "")
        }
    return timetable


def merge_timetable(existing_timetable, new_timetable):
    merged = {
        day: dict(periods)
        for day, periods in (existing_timetable or {}).items()
    }

    for day, periods in (new_timetable or {}).items():
        merged.setdefault(day, {})
        merged[day].update(periods)

    return merged


def get_period_from_times(start, end):
    normalized_start = str(start).strip() if start is not None else ""
    normalized_end = str(end).strip() if end is not None else ""
    for period, times in TIME_SLOTS.items():
        if times["start"] == normalized_start and times["end"] == normalized_end:
            return period
    return None


def normalize_class_type(value):
    class_type = str(value).strip().lower() if value is not None else ""
    return class_type or "theory"


def normalize_day(day):
    normalized_day = str(day).strip()
    if not normalized_day:
        raise HTTPException(400, "day is required")
    normalized_day = normalized_day.capitalize()
    if normalized_day not in VALID_DAYS:
        raise HTTPException(400, f"Invalid day: {day}")
    return normalized_day


def normalize_faculty_code(value):
    return str(value).strip().upper() if value is not None else ""


def find_faculty_by_code(faculty_code):
    normalized_code = normalize_faculty_code(faculty_code)
    if not normalized_code:
        return None

    faculty = db.faculty.find_one({"faculty_code": normalized_code})
    if faculty:
        return faculty

    return db.faculty.find_one({
        "faculty_code": {
            "$regex": f"^{normalized_code}$",
            "$options": "i"
        }
    })


def normalize_slot(slot):
    if not slot.get("day"):
        raise HTTPException(400, "Each slot must include day")

    period = slot.get("period")
    if period not in [None, ""]:
        try:
            period = int(period)
        except (TypeError, ValueError):
            raise HTTPException(400, f"Invalid period: {slot.get('period')}")

    if period is None:
        period = get_period_from_times(slot.get("start"), slot.get("end"))

    if period not in TIME_SLOTS:
        start = slot.get("start")
        end = slot.get("end")
        raise HTTPException(400, f"Invalid slot timing: {start} - {end}")

    return {
        "day": normalize_day(slot["day"]),
        "period": period,
        "subject": str(slot.get("subject", "")).strip(),
        "section": str(slot.get("section", "")).strip(),
        "class_type": normalize_class_type(slot.get("class_type") or slot.get("type")),
        "room": str(slot.get("room", "")).strip()
    }


def normalize_slots(slots):
    if not slots:
        raise HTTPException(400, "slots are required")
    return [normalize_slot(slot) for slot in slots]


def flatten_timetable(timetable):
    slots = []
    for day, periods in (timetable or {}).items():
        for period_key, period_data in periods.items():
            try:
                period = int(period_key)
            except (TypeError, ValueError):
                continue
            if period not in TIME_SLOTS:
                continue
            time_info = TIME_SLOTS[period]
            slots.append({
                "day": day,
                "period": period,
                "start": time_info["start"],
                "end": time_info["end"],
                "subject": period_data.get("subject", ""),
                "section": period_data.get("section", ""),
                "class_type": period_data.get("type", "theory"),
                "room": period_data.get("room", ""),
                "type": "class"
            })
    return sorted(slots, key=lambda slot: (slot["day"], slot["period"]))


def is_same_day(first, second):
    return first.date() == second.date()


def is_holiday(now):
    date_variants = [
        now.strftime("%Y-%m-%d"),
        now.strftime("%d-%m-%Y"),
        now.strftime("%Y/%m/%d"),
    ]
    return db.holidays.find_one({
        "$and": [
            {
                "$or": [
                    {"active": {"$exists": False}},
                    {"active": True}
                ]
            },
            {
                "$or": [
                    {"date": {"$in": date_variants}},
                    {"holiday_date": {"$in": date_variants}},
                ]
            }
        ]
    }) is not None


def get_manual_status_for_today(faculty, now):
    manual_status = faculty.get("manual_status")
    last_scan_at = faculty.get("last_scan_at")

    if not manual_status or not isinstance(last_scan_at, datetime):
        return None

    if not is_same_day(last_scan_at, now):
        return None

    return manual_status


def has_active_doubt_session(faculty):
    faculty_id = str(faculty.get("_id", ""))
    if not faculty_id:
        return False

    return db.doubts.count_documents({
        "faculty_id": faculty_id,
        "status": "active"
    }) > 0


def get_reserved_periods_for_day(faculty, target_date):
    faculty_id = str(faculty.get("_id", ""))
    if not faculty_id:
        return set()

    day_start = datetime.combine(target_date, time.min)
    day_end = day_start + timedelta(days=1)

    reserved = set()
    scheduled_doubts = db.doubts.find(
        {
            "faculty_id": faculty_id,
            "status": {"$in": ["scheduled", "active"]},
            "scheduled_for": {"$gte": day_start, "$lt": day_end},
            "scheduled_period": {"$exists": True}
        },
        {"scheduled_period": 1}
    )

    for doubt in scheduled_doubts:
        try:
            reserved.add(int(doubt.get("scheduled_period")))
        except (TypeError, ValueError):
            continue

    return reserved


def build_slot_payload(slot_date, period):
    slot_meta = TIME_SLOTS[period]
    return {
        "date": slot_date.isoformat(),
        "day": slot_date.strftime("%A"),
        "period": period,
        "start": slot_meta["start"],
        "end": slot_meta["end"],
        "label": f"{slot_date.strftime('%A')} | {slot_meta['start']} - {slot_meta['end']}",
    }


def get_upcoming_free_slots(faculty, timetable, start_dt=None, limit=5):
    start_dt = start_dt or datetime.now()
    upcoming = []

    for offset in range(0, 10):
        slot_date = (start_dt + timedelta(days=offset)).date()
        day_name = slot_date.strftime("%A")

        if day_name not in WORKING_DAYS:
            continue

        if is_holiday(datetime.combine(slot_date, time(hour=12))):
            continue

        day_schedule = timetable.get(day_name, {}) if timetable else {}
        busy_periods = set()
        for period_key in day_schedule.keys():
            try:
                busy_periods.add(int(period_key))
            except (TypeError, ValueError):
                continue

        reserved_periods = get_reserved_periods_for_day(faculty, slot_date)

        for period, slot_meta in TIME_SLOTS.items():
            slot_start = datetime.combine(
                slot_date,
                time(*map(int, slot_meta["start"].split(":")))
            )
            if slot_start < start_dt:
                continue
            if period in busy_periods or period in reserved_periods:
                continue

            upcoming.append(build_slot_payload(slot_date, period))
            if len(upcoming) >= limit:
                return upcoming

    return upcoming


def group_slots_by_day(slots):
    weekly = {day: [] for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]}
    for slot in slots:
        weekly.setdefault(slot["day"], [])
        weekly[slot["day"]].append(slot)

    for day_slots in weekly.values():
        day_slots.sort(key=lambda slot: slot["period"])

    return weekly


def save_faculty_timetable(filter_query, slots, merge=False):
    timetable = build_timetable_from_slots(slots)
    if merge:
        faculty = db.faculty.find_one(filter_query, {"timetable": 1})
        existing_timetable = faculty.get("timetable", {}) if faculty else {}
        timetable = merge_timetable(existing_timetable, timetable)

    result = db.faculty.update_one(
        filter_query,
        {"$set": {"timetable": timetable, "updated_at": datetime.utcnow()}}
    )
    return timetable, result


def get_status_from_timetable(faculty, timetable):
    now_dt = datetime.now()
    day = DAY_MAP.get(now_dt.weekday(), "Monday")
    current_period = get_current_period()
    now = now_dt.strftime("%H:%M")
    manual_status = get_manual_status_for_today(faculty, now_dt)
    reserved_periods = get_reserved_periods_for_day(faculty, now_dt.date())
    upcoming_slots = get_upcoming_free_slots(faculty, timetable, now_dt, limit=5)
    today_free_slots = [slot for slot in upcoming_slots if slot["day"] == day]

    if day not in WORKING_DAYS or is_holiday(now_dt):
        return "holiday", "Holiday - Faculty unavailable", today_free_slots

    if manual_status == "left" or now >= "17:00":
        return "left", "Faculty has left for the day", today_free_slots

    if manual_status != "available":
        return "not_checked_in", "Faculty has not checked in today", today_free_slots

    day_schedule = timetable.get(day, {}) if timetable else {}
    busy_periods = [int(k) for k in day_schedule.keys()]

    if has_active_doubt_session(faculty):
        return "busy", "Attending a student doubt", today_free_slots
    elif current_period and current_period in reserved_periods:
        return "busy", "Scheduled doubt session", today_free_slots
    elif current_period and current_period in busy_periods:
        current_slot = day_schedule.get(str(current_period), {})
        current_subject = current_slot.get("subject") or faculty.get("subject", "")
        return "in_class", f"In class - {current_subject}", today_free_slots
    elif now < "09:10":
        return "available", "Checked in - Available before first class", today_free_slots
    elif "12:30" <= now < "13:30":
        return "lunch", "Lunch break", today_free_slots
    else:
        return "available", "Checked in - Available for doubt sessions", today_free_slots


def get_average_resolution_minutes(faculty_id, limit=20):
    completed = list(db.doubts.find({
        "faculty_id": faculty_id,
        "status": "completed",
        "accepted_at": {"$exists": True},
        "completed_at": {"$exists": True}
    }).sort("completed_at", -1).limit(limit))

    durations = []
    for doubt in completed:
        accepted_at = doubt.get("accepted_at")
        completed_at = doubt.get("completed_at")
        if isinstance(accepted_at, datetime) and isinstance(completed_at, datetime):
            diff = int((completed_at - accepted_at).total_seconds() // 60)
            if 1 <= diff <= 120:
                durations.append(diff)

    if not durations:
        return DEFAULT_RESOLUTION_MINUTES

    return round(sum(durations) / len(durations))


@router.get("/all")
def get_all_timetables(authorization: str = Header(...)):
    verify_user(authorization, ["admin"])
    faculty_list = list(db.faculty.find({}))
    result = []

    for faculty in faculty_list:
        timetable = faculty.get("timetable", {})
        slots = flatten_timetable(timetable)
        result.append({
            "faculty_code": faculty.get("faculty_code"),
            "faculty_name": faculty.get("name"),
            "email": faculty.get("email"),
            "subject": faculty.get("subject"),
            "timetable": timetable,
            "slots": slots,
            "updated_at": faculty.get("updated_at")
        })

    return {"timetables": result}


@router.post("/upload")
def upload_timetable(data: dict, authorization: str = Header(...)):
    verify_user(authorization, ["admin"])

    faculty_code = normalize_faculty_code(data.get("faculty_code"))
    slots = data.get("slots", [])

    if not faculty_code:
        raise HTTPException(400, "faculty_code is required")

    faculty = find_faculty_by_code(faculty_code)
    if not faculty:
        raise HTTPException(404, "Faculty not found")

    normalized = normalize_slots(slots)
    timetable, result = save_faculty_timetable({"_id": faculty["_id"]}, normalized, merge=True)

    return {
        "message": f"Timetable saved for {faculty.get('faculty_code', faculty_code)}",
        "slots": flatten_timetable(timetable)
    }


@router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...), authorization: str = Header(...)):
    verify_user(authorization, ["admin"])

    if not file.filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(400, "Upload a valid Excel file")

    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    # Supported headers:
    # faculty_code | day | period | subject | section | class_type | room
    # faculty_code | day | start | end | subject | section | class_type | room
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Excel file is empty")

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    required = ["faculty_code", "day", "subject"]
    for col in required:
        if col not in headers:
            raise HTTPException(400, f"Missing required column: {col}")
    has_period = "period" in headers
    has_times = "start" in headers and "end" in headers
    if not has_period and not has_times:
        raise HTTPException(400, "Excel must include either period or start and end columns")

    idx = {h: i for i, h in enumerate(headers)}
    grouped = {}

    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue

        faculty_code = str(row[idx["faculty_code"]]).strip()
        if not faculty_code:
            continue

        subject = str(row[idx["subject"]]).strip() if row[idx["subject"]] is not None else ""
        section = str(row[idx["section"]]).strip() if "section" in idx and row[idx["section"]] else ""
        class_type = str(row[idx["class_type"]]).strip() if "class_type" in idx and row[idx["class_type"]] else "theory"
        room = str(row[idx["room"]]).strip() if "room" in idx and row[idx["room"]] else ""

        slot_payload = {
            "day": row[idx["day"]],
            "subject": subject,
            "section": section,
            "class_type": class_type,
            "room": room
        }

        if has_period:
            slot_payload["period"] = row[idx["period"]]
        else:
            slot_payload["start"] = row[idx["start"]]
            slot_payload["end"] = row[idx["end"]]

        normalized_slot = normalize_slot(slot_payload)

        grouped.setdefault(faculty_code, [])
        grouped[faculty_code].append(normalized_slot)

    saved = []
    for faculty_code, slots in grouped.items():
        _, result = save_faculty_timetable({"faculty_code": faculty_code}, slots)
        if result.matched_count:
            saved.append(faculty_code)

    return {"message": "Excel upload processed", "saved": saved}


@router.delete("/delete/{faculty_code}")
def delete_timetable(faculty_code: str, authorization: str = Header(...)):
    verify_user(authorization, ["admin"])
    result = db.faculty.update_one(
        {"faculty_code": faculty_code},
        {"$unset": {"timetable": ""}}
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Faculty not found")
    return {"message": f"Timetable deleted for {faculty_code}"}


@router.post("/faculty-upload")
def faculty_upload_timetable(data: dict, authorization: str = Header(...)):
    user = verify_user(authorization, ["faculty"])
    normalized = normalize_slots(data.get("slots", []))
    timetable, result = save_faculty_timetable({"_id": ObjectId(user["id"])}, normalized)
    if result.matched_count == 0:
        raise HTTPException(404, "Faculty not found")
    return {
        "message": "Your timetable was saved successfully",
        "slots": flatten_timetable(timetable)
    }


@router.get("/faculty-status/{faculty_code}")
def get_faculty_status(faculty_code: str):
    faculty = find_faculty_by_code(faculty_code)
    if not faculty:
        raise HTTPException(404, "Faculty not found")

    timetable = faculty.get("timetable", {})
    status, message, free_slots = get_status_from_timetable(faculty, timetable)
    next_free_slot = get_upcoming_free_slots(faculty, timetable, datetime.now(), limit=1)

    return {
        "faculty_code": faculty.get("faculty_code", faculty_code),
        "faculty_name": faculty["name"],
        "subject": faculty.get("subject", ""),
        "cabin": faculty.get("cabin", ""),
        "email": faculty.get("email", ""),
        "status": status,
        "message": message,
        "current_day": get_current_day(),
        "free_slots_today": free_slots,
        "next_free_slot": next_free_slot[0] if next_free_slot else None,
    }


@router.get("/best-slot/{faculty_code}")
def get_best_slot(faculty_code: str):
    faculty = db.faculty.find_one({"faculty_code": faculty_code})
    if not faculty:
        raise HTTPException(404, "Faculty not found")

    timetable = faculty.get("timetable", {})
    day = get_current_day()
    now = datetime.now().strftime("%H:%M")
    day_schedule = timetable.get(day, {})
    busy_periods = [int(k) for k in day_schedule.keys()]

    free_slots = []
    for period, times in TIME_SLOTS.items():
        if period not in busy_periods and times["start"] >= now:
            free_slots.append({
                "period": period,
                "start": times["start"],
                "end": times["end"],
            })

    if free_slots:
        return {"best_slot": free_slots[0], "day": day, "message": f"Next available: {free_slots[0]['start']} today"}

    return {"message": "No slots available today", "best_slot": None}


@router.get("/all-faculty-status")
def get_all_faculty_status():
    faculties = list(db.faculty.find({}))
    result = []

    for faculty in faculties:
        code = faculty.get("faculty_code", "")
        timetable = faculty.get("timetable", {})
        status, message, free_slots = get_status_from_timetable(faculty, timetable)
        next_free_slots = get_upcoming_free_slots(faculty, timetable, datetime.now(), limit=1)
        next_free_slot = next_free_slots[0] if next_free_slots else None

        queue_count = db.doubts.count_documents({
            "faculty_id": str(faculty["_id"]),
            "status": "pending",
        })
        average_resolution_minutes = get_average_resolution_minutes(str(faculty["_id"]))
        estimated_wait = calculate_wait_time(
            str(faculty["_id"]),
            max(1, queue_count + 1),
            db,
            {
                "status": status,
                "free_slots_today": free_slots,
                "next_free_slot": next_free_slot,
            }
        )

        result.append({
            "_id": str(faculty["_id"]),
            "faculty_code": code,
            "faculty_name": faculty["name"],
            "subject": faculty.get("subject", ""),
            "cabin": faculty.get("cabin", ""),
            "block": faculty.get("block", ""),
            "email": faculty.get("email", f"{code.lower()}@kiet.edu"),
            "status": status,
            "message": message,
            "free_slots_today": free_slots,
            "queue_count": queue_count,
            "average_resolution_minutes": average_resolution_minutes,
            "estimated_wait": estimated_wait.get("estimated_wait"),
            "estimated_wait_minutes": estimated_wait.get("estimated_wait_minutes"),
            "expected_free_time": estimated_wait.get("expected_free_time"),
            "next_free_slot": next_free_slot,
        })

    return {"faculty": result}


@router.get("/my-schedule")
def get_my_schedule(authorization: str = Header(...)):
    user = verify_user(authorization, ["faculty"])
    faculty = db.faculty.find_one({"_id": ObjectId(user["id"])})
    if not faculty:
        raise HTTPException(404, "Faculty not found")

    timetable = faculty.get("timetable", {})
    weekly_slots = group_slots_by_day(flatten_timetable(timetable))
    day = get_current_day()
    slots = []
    for period in range(1, 9):
        period_data = timetable.get(day, {}).get(str(period), None)
        time_info = TIME_SLOTS[period]
        slot = {
            "period": period,
            "start": time_info["start"],
            "end": time_info["end"],
        }
        if period_data:
            slot.update({
                "type": "class",
                "subject": period_data.get("subject", ""),
                "section": period_data.get("section", ""),
                "class_type": period_data.get("type", "theory"),
                "room": period_data.get("room", ""),
            })
        else:
            slot["type"] = "free"
        slots.append(slot)

    return {
        "day": day,
        "faculty_name": faculty["name"],
        "slots": slots,
        "weekly": weekly_slots,
        "timetable": timetable
    }
